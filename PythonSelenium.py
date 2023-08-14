import time
import datetime
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver.get("https://google.com")
driver.maximize_window()

day = datetime.datetime.now().strftime("%A")

LS = load_workbook('data.xlsx')
ES = LS.active

source = LS[day]
i=3
longest = []
shortest = []
for cell in source['C']:
    d = source.cell(row= i, column=3)
    driver.find_element(By.NAME, "q").send_keys(d.value)
    time.sleep(2)

    s_id = driver.find_element(By.ID, "Alh6id")
    element = s_id.find_elements(By.CLASS_NAME, "wM6W7d")
    word = []
    for adds in element:
        word.append(adds.text)

    worrd = max(word, key=len)
    wornd = min(word, key=len)

    longest.append(worrd)
    shortest.append(wornd)
    driver.find_element(By.NAME, "q").clear()
    if (i== len(source['C'])):
        break
    i=i+1
driver.close()


j=3
for cell in source['C']:
    source.cell(row=j, column=4, value = longest[j-3])
    source.cell(row=j, column=5, value = shortest[j-3])

    if (j== len(source['C'])):
        break
    j=j+1

LS.save('data.xlsx')